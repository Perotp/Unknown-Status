# import openpyxl as op
# from openpyxl import load_workbook  
# wb = load_workbook('C:\\Users\\10383563\\Downloads\\Proposal-Project Numbers1.xlsx')  

# op.Workbook.

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

filepath = 'C:\\Users\\10383563\\Downloads\\Proposal-Project Numbers1.xlsx'
wb = load_workbook(filepath)
wb.close
sheet = wb['Sheet1']
# data = sheet.values
# cols = next(data)


# df = pd.DataFrame(data, columns=cols)
required_cols = [0, 1, 2, 3, 4, 5, 6, 7]
df = pd.read_excel(filepath, sheet_name = 'Sheet1', header=0, usecols = required_cols)

print(df)

sorted_df = df.sort_values(by=['Date'])

print(sorted_df)

rowcnt = len(df)
# get last project number


info = sorted_df['Seq'].values[rowcnt-1]
print(info)
newinfo = info+1
print(newinfo)

# Add entry to existing file
Area = '2200'
Seq = newinfo
Proj = "Some Project"
Location = "Some Location"
Date = "2024-07-09"
Manager = "PWP"
Admin = "HE"
threel  =   "ADD"

newdat = [Area, Seq, Proj, Location, Date, Manager, Admin, threel]
sheet.append(newdat)
wb.save(filepath)

# newinfo = 

# sheet.append(newinfo)
