# Python program to read an excel file
import openpyxl 
#import pyautogui
# Give the location of the file
# path = "C:\\Users\\10383563\\Downloads\\Proposal-Project Numbers1.xlsx"
# path = "C:\\Users\\10383563\\Downloads\\Proposal Setup Sheet POLAT-0.009"
# To open the workbook 
# workbook object is created
# wb = openpyxl.load_workbook(path)
 
# Get workbook active sheet object
# from the active attribute
# sheet = wb.active
# col = "B"
# last_row = None
# for cell in sheet[col]:
#     if cell.value:
#         last_row = cell.row

# # print(last_row)
# mcell = col+str(last_row)
# mycell=sheet[mcell].value
# print(mycell)
# scell = "A2"
# ecol = "I"
# mran = scell+":"+ecol+str(last_row)
# mrange = sheet[mran]
# mrange = sheet["A2:I1015"]
# print(mran)

# # Sort the data by column A in ascending order
# sorted_data = sorted("A2:I1015", key=lambda x: "E")

# # Replace the original data with the sorted data
# for idx, row in enumerate(sorted_data, start=2):
#     for j, cell in enumerate(row, start=1):
#         sheet.cell(row=idx, column=j).value = cell.value


# mycell=sheet[mcell].value
# print(mycell)


path = "C:\\Users\\10383563\\Downloads\\Proposal Setup Sheet POLAT-0009.xlsx"
# To open the workbook 
# workbook object is created
wb = openpyxl.load_workbook(path)
wb.active = wb["Customers"]
ws = wb.active
col = "B"
last_row = None
for cell in ws[col]:
    if cell.value:
        last_row = cell.row

print(last_row)