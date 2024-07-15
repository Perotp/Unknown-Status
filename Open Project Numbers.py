from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from openpyxl.styles import NamedStyle



#   This section loads the Project Number file and sets it to sheet1
filepath = 'C:\\Users\\10383563\\Downloads\\Proposal-Project Numbers1.xlsx'


# This section extracts the data from the workbook and makes a dataframe
required_cols = [0, 1, 2, 3, 4, 5, 6, 7]
df = pd.read_excel(filepath, sheet_name = 'Sheet1', header=0, usecols = required_cols)
# print(df)

#  This section sorts the dataFrame by the title DATE
sorted_df = df.sort_values(by=['Date'])
# print(sorted_df)
rowcnt = len(df)

# get last project number
info = sorted_df['Seq'].values[rowcnt-1]
# print(info)
newinfo = info+1
print(newinfo)

# Add entry to existing file
#  The values here except for newinfo will come from the GUI
Area = '2200'
Seq = newinfo
Proj = "Some Project"
Location = "Some Location"
Date = "2024-07-09"
Manager = "PWP"
Admin = "HE"
threel  =   "GCC"

wb = load_workbook(filepath)
sheet = wb['Sheet1']
print(rowcnt)
mycell = sheet.cell(row=rowcnt, column=1) 
mycell = Area

print(mycell)
rowcnt=rowcnt + 2
# mycell = sheet.cell(row=rowcnt, column=1) 
# mycell.value = Area
# mycell = sheet.cell(row=rowcnt, column=2) 
# mycell.value = Seq
# mycell = sheet.cell(row=rowcnt, column=3) 
# mycell.value  = Proj
# mycell = sheet.cell(row=rowcnt, column=4) 
# mycell.value  = Location
# mycell = sheet.cell(row=rowcnt, column=5) 
# mycell.value  = Date
# mycell = sheet.cell(row=rowcnt, column=6) 
# mycell.value  = Manager
# mycell = sheet.cell(row=rowcnt, column=7) 
# mycell.value  = Admin
# mycell = sheet.cell(row=rowcnt, column=8) 
# mycell.value  = threel

date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')


sheet.cell(row=rowcnt, column=1, value = Area)
sheet.cell(row=rowcnt, column=2, value = Seq)

sheet.cell(row=rowcnt, column=3, value = Proj)
sheet.cell(row=rowcnt, column=4, value = Location)
sheet.cell(row=rowcnt, column=5, value = Date)

mycell = sheet.cell(row=rowcnt, column=5)

# print(mycell)

mycell.style = date_style

sheet.cell(row=rowcnt, column=6,  value = Manager)
sheet.cell(row=rowcnt, column=7,  value = Admin)
sheet.cell(row=rowcnt, column=8,  value = threel)

wb.save(filepath)
